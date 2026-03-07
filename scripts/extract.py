"""
Extract members data from the Excel file into a clean JSON file.

Usage:
    python3 scripts/extract.py "Members Data - Apps.xlsx"
    python3 scripts/extract.py path/to/other.xlsx

Output: scripts/data.json
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


def parse_national_id(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return str(int(val))
    return str(val).strip()


def split_teams(raw: str) -> list[str]:
    """Split 'CS (Computer Society), RAS (Robotics & Automation Society)' into individual teams."""
    import re
    return [t.strip() for t in re.split(r",\s*(?=[A-Z])", raw) if t.strip()]


def extract_sheet(ws, name_col: str, team_col: str) -> list[dict]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        name = str(data.get(name_col) or "").strip()
        email = str(data.get("Email Address") or "").strip()
        if not name or not email:
            continue
        raw_team = str(data.get(team_col) or "").strip()
        rows.append({
            "name": name,
            "email": email,
            "phone": str(data.get("Phone Number") or "").strip(),
            "national_id": parse_national_id(data.get("National ID Number")),
            "position": str(data.get("Current Position") or "").strip(),
            "teams": split_teams(raw_team) if raw_team else [],
        })
    return rows


def deduplicate(members: list[dict]) -> list[dict]:
    seen_email = set()
    seen_nid = set()
    unique = []
    for m in members:
        if m["email"] in seen_email:
            continue
        if m["national_id"] and m["national_id"] in seen_nid:
            continue
        seen_email.add(m["email"])
        if m["national_id"]:
            seen_nid.add(m["national_id"])
        unique.append(m)
    return unique


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/extract.py <excel-file>")
        sys.exit(1)

    filepath = sys.argv[1]
    wb = openpyxl.load_workbook(filepath)

    members = []

    # Sheet configs: (sheet_name, name_column, team_column)
    sheet_configs = [
        ("Form Responses 1", "Quad name (English)", "Chapters"),
        ("Talent & Tech", "Name", "Teams"),
    ]

    for sheet_name, name_col, team_col in sheet_configs:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found, skipping")
            continue
        rows = extract_sheet(wb[sheet_name], name_col, team_col)
        members.extend(rows)
        print(f"Extracted {len(rows)} rows from '{sheet_name}'")

    unique = deduplicate(members)
    teams = sorted(set(t for m in unique for t in m["teams"]))

    output = {"teams": teams, "members": unique}
    out_path = Path(__file__).parent / "data.json"
    out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False))

    print(f"\nOutput: {out_path}")
    print(f"Teams: {len(teams)}, Members: {len(unique)}")


if __name__ == "__main__":
    main()
